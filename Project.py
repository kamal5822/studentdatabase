import tkinter
from tkinter import *
from tkinter import ttk
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import messagebox

def AddNew():
    AddNewFrame.pack(fill='both',expand=1)
    HomeFrame.pack_forget()
    SearchFrame.pack_forget()

def Home():
    HomeFrame.pack(fill='both',expand=1)
    AddNewFrame.pack_forget()
    SearchFrame.pack_forget()

def Search():
    SearchFrame.pack(fill='both',expand=1)
    HomeFrame.pack_forget()
    AddNewFrame.pack_forget()

def Reset():
    AdmissionNo.set("")
    ChildId.set("")
    StudentName.set("")
    Gender.set("")
    DateOfBirth.set("")
    DateOfJoining.set("")
    AadharNo.set("")
    FatherName.set("")
    MotherName.set("")
    GuardianName.set("")
    MobileNo.set("")
    Residence.set("")
    Caste.set("")
    SubCaste.set("")
    RationCardNo.set("")
    BankAccountNo.set("")
    IfscCode.set("")
    BankName.set("")
    Class.set("")

def Exit():
    win.destroy()

def AddData():
    wb = Workbook()
    wb = load_workbook('database.xlsx')
    ws = wb.active
    lrow = [AdmissionNo.get(),ChildId.get(),StudentName.get(),Gender.get(),DateOfBirth.get(),DateOfJoining.get()
            ,AadharNo.get(),FatherName.get(),MotherName.get(),GuardianName.get(),MobileNo.get(),Residence.get()
            ,Caste.get(),SubCaste.get(),RationCardNo.get(),BankAccountNo.get(),IfscCode.get(),BankName.get()
            ,Class.get()]
    ws.append(lrow)
    wb.save('database.xlsx')
    Reset()

def SearchData():
    wb = Workbook()
    wb = load_workbook('database.xlsx')
    ws = wb.active
    if AdmissionNo.get() != '':
        for cell in ws['A']:
            if str(cell.value) == AdmissionNo.get():
                AdmissionNo.set(ws.cell(row=cell.row,column=1).value)
                ChildId.set(ws.cell(row=cell.row,column=2).value)
                StudentName.set(ws.cell(row=cell.row,column=3).value)
                Gender.set(ws.cell(row=cell.row,column=4).value)
                DateOfBirth.set(ws.cell(row=cell.row,column=5).value)
                DateOfJoining.set(ws.cell(row=cell.row,column=6).value)
                AadharNo.set(ws.cell(row=cell.row,column=7).value)
                FatherName.set(ws.cell(row=cell.row,column=8).value)
                MotherName.set(ws.cell(row=cell.row,column=9).value)
                GuardianName.set(ws.cell(row=cell.row,column=10).value)
                MobileNo.set(ws.cell(row=cell.row,column=11).value)
                Residence.set(ws.cell(row=cell.row,column=12).value)
                Caste.set(ws.cell(row=cell.row,column=13).value)
                SubCaste.set(ws.cell(row=cell.row,column=14).value)
                RationCardNo.set(ws.cell(row=cell.row,column=15).value)
                BankAccountNo.set(ws.cell(row=cell.row,column=16).value)
                IfscCode.set(ws.cell(row=cell.row,column=17).value)
                BankName.set(ws.cell(row=cell.row,column=18).value)
                Class.set(ws.cell(row=cell.row,column=19).value)
                break;
    elif StudentName.get() != '':
        for cell in ws['C']:
            if str(cell.value) == StudentName.get():
                AdmissionNo.set(ws.cell(row=cell.row,column=1).value)
                ChildId.set(ws.cell(row=cell.row,column=2).value)
                StudentName.set(ws.cell(row=cell.row,column=3).value)
                Gender.set(ws.cell(row=cell.row,column=4).value)
                DateOfBirth.set(ws.cell(row=cell.row,column=5).value)
                DateOfJoining.set(ws.cell(row=cell.row,column=6).value)
                AadharNo.set(ws.cell(row=cell.row,column=7).value)
                FatherName.set(ws.cell(row=cell.row,column=8).value)
                MotherName.set(ws.cell(row=cell.row,column=9).value)
                GuardianName.set(ws.cell(row=cell.row,column=10).value)
                MobileNo.set(ws.cell(row=cell.row,column=11).value)
                Residence.set(ws.cell(row=cell.row,column=12).value)
                Caste.set(ws.cell(row=cell.row,column=13).value)
                SubCaste.set(ws.cell(row=cell.row,column=14).value)
                RationCardNo.set(ws.cell(row=cell.row,column=15).value)
                BankAccountNo.set(ws.cell(row=cell.row,column=16).value)
                IfscCode.set(ws.cell(row=cell.row,column=17).value)
                BankName.set(ws.cell(row=cell.row,column=18).value)
                Class.set(ws.cell(row=cell.row,column=19).value)
                break;
    else:
            Reset()

def DeleteData():
    wb = Workbook()
    wb = load_workbook('database.xlsx')
    ws = wb.active
    if AdmissionNo.get() != '':
        for cell in ws['A']:
            if str(cell.value) == AdmissionNo.get():
                AdmissionNo.set(ws.cell(row=cell.row,column=1).value)
                ChildId.set(ws.cell(row=cell.row,column=2).value)
                StudentName.set(ws.cell(row=cell.row,column=3).value)
                Gender.set(ws.cell(row=cell.row,column=4).value)
                DateOfBirth.set(ws.cell(row=cell.row,column=5).value)
                DateOfJoining.set(ws.cell(row=cell.row,column=6).value)
                AadharNo.set(ws.cell(row=cell.row,column=7).value)
                FatherName.set(ws.cell(row=cell.row,column=8).value)
                MotherName.set(ws.cell(row=cell.row,column=9).value)
                GuardianName.set(ws.cell(row=cell.row,column=10).value)
                MobileNo.set(ws.cell(row=cell.row,column=11).value)
                Residence.set(ws.cell(row=cell.row,column=12).value)
                Caste.set(ws.cell(row=cell.row,column=13).value)
                SubCaste.set(ws.cell(row=cell.row,column=14).value)
                RationCardNo.set(ws.cell(row=cell.row,column=15).value)
                BankAccountNo.set(ws.cell(row=cell.row,column=16).value)
                IfscCode.set(ws.cell(row=cell.row,column=17).value)
                BankName.set(ws.cell(row=cell.row,column=18).value)
                Class.set(ws.cell(row=cell.row,column=19).value)
                ws.delete_rows(cell.row,1)
                wb.save('database.xlsx')
                break;
    elif StudentName.get() != '':
        for cell in ws['C']:
            if str(cell.value) == StudentName.get():
                AdmissionNo.set(ws.cell(row=cell.row,column=1).value)
                ChildId.set(ws.cell(row=cell.row,column=2).value)
                StudentName.set(ws.cell(row=cell.row,column=3).value)
                Gender.set(ws.cell(row=cell.row,column=4).value)
                DateOfBirth.set(ws.cell(row=cell.row,column=5).value)
                DateOfJoining.set(ws.cell(row=cell.row,column=6).value)
                AadharNo.set(ws.cell(row=cell.row,column=7).value)
                FatherName.set(ws.cell(row=cell.row,column=8).value)
                MotherName.set(ws.cell(row=cell.row,column=9).value)
                GuardianName.set(ws.cell(row=cell.row,column=10).value)
                MobileNo.set(ws.cell(row=cell.row,column=11).value)
                Residence.set(ws.cell(row=cell.row,column=12).value)
                Caste.set(ws.cell(row=cell.row,column=13).value)
                SubCaste.set(ws.cell(row=cell.row,column=14).value)
                RationCardNo.set(ws.cell(row=cell.row,column=15).value)
                BankAccountNo.set(ws.cell(row=cell.row,column=16).value)
                IfscCode.set(ws.cell(row=cell.row,column=17).value)
                BankName.set(ws.cell(row=cell.row,column=18).value)
                Class.set(ws.cell(row=cell.row,column=19).value)
                ws.delete_rows(cell.row,1)
                wb.save('database.xlsx')
                break;
    else:
        Reset()

def UpdateData():
    wb = Workbook()
    wb = load_workbook('database.xlsx')
    ws = wb.active
    if AdmissionNo.get() != '':
        for cell in ws['A']:
            if str(cell.value) == AdmissionNo.get():
                ws.cell(row = cell.row,column = 1).value = AdmissionNo.get()
                ws.cell(row = cell.row,column = 2).value = ChildId.get()
                ws.cell(row = cell.row,column = 3).value = StudentName.get()
                ws.cell(row = cell.row,column = 4).value = Gender.get()
                ws.cell(row = cell.row,column = 5).value = DateOfBirth.get()
                ws.cell(row = cell.row,column = 6).value = DateOfJoining.get()
                ws.cell(row = cell.row,column = 7).value = AadharNo.get()
                ws.cell(row = cell.row,column = 8).value = FatherName.get()
                ws.cell(row = cell.row,column = 9).value = MotherName.get()
                ws.cell(row = cell.row,column = 10).value = GuardianName.get()
                ws.cell(row = cell.row,column = 11).value = MobileNo.get()
                ws.cell(row = cell.row,column = 12).value = Residence.get()
                ws.cell(row = cell.row,column = 13).value = Caste.get()
                ws.cell(row = cell.row,column = 14).value = SubCaste.get()
                ws.cell(row = cell.row,column = 15).value = RationCardNo.get()
                ws.cell(row = cell.row,column = 16).value = BankAccountNo.get()
                ws.cell(row = cell.row,column = 17).value = IfscCode.get()
                ws.cell(row = cell.row,column = 18).value = BankName.get()
                ws.cell(row = cell.row,column = 19).value = Class.get()
                wb.save('database.xlsx')
                Reset()
                break;
    elif StudentName.get() != '':
        for cell in ws['C']:
            if str(cell.value) == StudentName.get():
                ws.cell(row = cell.row,column = 1).value = AdmissionNo.get()
                ws.cell(row = cell.row,column = 2).value = ChildId.get()
                ws.cell(row = cell.row,column = 3).value = StudentName.get()
                ws.cell(row = cell.row,column = 4).value = Gender.get()
                ws.cell(row = cell.row,column = 5).value = DateOfBirth.get()
                ws.cell(row = cell.row,column = 6).value = DateOfJoining.get()
                ws.cell(row = cell.row,column = 7).value = AadharNo.get()
                ws.cell(row = cell.row,column = 8).value = FatherName.get()
                ws.cell(row = cell.row,column = 9).value = MotherName.get()
                ws.cell(row = cell.row,column = 10).value = GuardianName.get()
                ws.cell(row = cell.row,column = 11).value = MobileNo.get()
                ws.cell(row = cell.row,column = 12).value = Residence.get()
                ws.cell(row = cell.row,column = 13).value = Caste.get()
                ws.cell(row = cell.row,column = 14).value = SubCaste.get()
                ws.cell(row = cell.row,column = 15).value = RationCardNo.get()
                ws.cell(row = cell.row,column = 16).value = BankAccountNo.get()
                ws.cell(row = cell.row,column = 17).value = IfscCode.get()
                ws.cell(row = cell.row,column = 18).value = BankName.get()
                ws.cell(row = cell.row,column = 19).value = Class.get()
                wb.save('database.xlsx')
                Reset()
                break;
    else:
        Reset()

win = Tk()

win.geometry("1400x800")

MainFrame = Frame(win,bd=5,width=1390,height=790,relief=RIDGE,bg="cadet blue")
MainFrame.grid()

TitleFrame = Frame(MainFrame,bd=5,width=1385,height = 90,relief=RIDGE)
TitleFrame.grid(row=0,column=0)

TopFrame = Frame(MainFrame,bd=5,width=1380,height = 690,relief=RIDGE)
TopFrame.grid(row=1,column=0)

LeftFrame = Frame(TopFrame,bd=5,width=280,height = 690,relief=RIDGE)
LeftFrame.grid(row=0,column=0)

RightFrame = Frame(TopFrame,bd=5,width=1100,height = 690,relief=RIDGE)
RightFrame.grid(row=0,column=1)

#====================================================================================================================

AdmissionNo = StringVar()
ChildId = StringVar()
StudentName = StringVar()
Gender = StringVar()
DateOfBirth = StringVar()
DateOfJoining = StringVar()
AadharNo = StringVar()
FatherName = StringVar()
MotherName = StringVar()
GuardianName = StringVar()
MobileNo = StringVar()
Residence = StringVar()
Caste = StringVar()
SubCaste = StringVar()
RationCardNo = StringVar()
BankAccountNo = StringVar()
IfscCode = StringVar()
BankName = StringVar()
Class = StringVar()

#====================================================================================================================

lblTitle = Label(TitleFrame,font = ('arial',42,'bold'),text="M.P.P.S Duppalapudi",bd=7)
lblTitle.grid(row=0,column=0,padx=130)

#====================================================================================================================

btnHome = Button(LeftFrame,pady=5,bd=5,font=('arial',16,'bold'),padx=5,height=2,
                   width=18,text='Home Page',command = Home).grid(row=0,column=0)
btnAddNew = Button(LeftFrame,pady=5,bd=5,font=('arial',16,'bold'),padx=5,height=2,
                   width=18,text='Add New Student',command= AddNew).grid(row=1,column=0)
btnSearch = Button(LeftFrame,pady=5,bd=5,font=('arial',16,'bold'),padx=5,height=2,
                   width=18,text='Search Student',command= Search).grid(row=2,column=0)
btnList = Button(LeftFrame,pady=5,bd=5,font=('arial',16,'bold'),padx=5,height=2,
                   width=18,text='Student Lists').grid(row=3,column=0)

btnAddNew = Button(LeftFrame,pady=5,bd=5,font=('arial',16,'bold'),padx=5,height=2,
                   width=18,text='Add New Student').grid(row=4,column=0)
btnSearch = Button(LeftFrame,pady=5,bd=5,font=('arial',16,'bold'),padx=5,height=2,
                   width=18,text='Search Student').grid(row=5,column=0)

btnAddNew = Button(LeftFrame,pady=5,bd=5,font=('arial',16,'bold'),padx=5,height=2,
                   width=18,text='Add New Student').grid(row=6,column=0)
btnExit = Button(LeftFrame,pady=5,bd=5,font=('arial',16,'bold'),padx=5,height=2,
                   width=18,text='Exit Application',command= Exit).grid(row=7,column=0)

#====================================================================================================================

HomeFrame = Frame(RightFrame,bd=0,width=1095,height=680,relief=RIDGE)

AddNewFrame = Frame(RightFrame,bd=0,width=1095,height=680,relief=RIDGE)
AddNewFrame1 = Frame(AddNewFrame,bd=5,width=1095,height=590,relief=RIDGE)
AddNewFrame1.grid(row=0,column=0)
AddNewFrame2 = Frame(AddNewFrame,bd=5,width=1095,height=80,relief=RIDGE)
AddNewFrame2.grid(row=1,column=0)

SearchFrame = Frame(RightFrame,bd=0,width=1095,height=680,relief=RIDGE)
SearchFrame1 = Frame(SearchFrame,bd=5,width=1095,height=80,relief=RIDGE)
SearchFrame1.grid(row=0,column=0)
SearchFrame2 = Frame(SearchFrame,bd=5,width=1095,height=500,relief=RIDGE)
SearchFrame2.grid(row=1,column=0)
SearchFrame3 = Frame(SearchFrame,bd=5,width=1095,height=80,relief=RIDGE)
SearchFrame3.grid(row=2,column=0)

HomeFrame.pack(fill='both',expand=1)

#====================================================================================================================

lblAdmissionNo = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Admission No",width=20,pady=10,)
lblAdmissionNo.grid(row=0,column=0)
txtAdmissionNo = Entry(AddNewFrame1,font=('arial',12,'bold'),width=36,justify='left',textvariable=AdmissionNo)
txtAdmissionNo.grid(row=0,column=1)

lblStudentName = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Student Name",width=20,pady=10)
lblStudentName.grid(row=0,column=2)
txtStudentName = Entry(AddNewFrame1,font=('arial',12,'bold'),width=36,justify='left',textvariable=StudentName)
txtStudentName.grid(row=0,column=3)

lblChildId = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Child Id",width=20,pady=10)
lblChildId.grid(row=1,column=0)
txtChildId = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=ChildId)
txtChildId.grid(row=1,column=1)


lblGender = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Gender",width=20,pady=10)
lblGender.grid(row=1,column=2)
cboGender = ttk.Combobox(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=Gender)
cboGender.grid(row=1,column=3)
cboGender['values'] = ('','Girl','Boy')
cboGender.current(0)

lblDateOfBirth = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Date Of Birth(dd/mm/yy)",width=20,pady=10)
lblDateOfBirth.grid(row=2,column=0)
txtDateOfBirth = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=DateOfBirth)
txtDateOfBirth.grid(row=2,column=1)

lblDateOfJoining = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Date Of Joining(dd/mm/yy)",width=20,pady=10)
lblDateOfJoining.grid(row=2,column=2)
txtDateOfJoining = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=DateOfJoining)
txtDateOfJoining.grid(row=2,column=3)

lblAadharNo = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Aadhar Card No",width=20,pady=10)
lblAadharNo.grid(row=3,column=0)
txtAadharNo = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=AadharNo)
txtAadharNo.grid(row=3,column=1)

lblFatherName = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Father Name",width=20,pady=10)
lblFatherName.grid(row=3,column=2)
txtFatherName = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=FatherName)
txtFatherName.grid(row=3,column=3)

lblMotherName = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Mother Name",width=20,pady=10)
lblMotherName.grid(row=4,column=0)
txtMotherName = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=MotherName)
txtMotherName.grid(row=4,column=1)

lblGuardianName = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Guardian Name",width=20,pady=10)
lblGuardianName.grid(row=4,column=2)
txtGuardianName = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=GuardianName)
txtGuardianName.grid(row=4,column=3)

lblMobileNo = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Mobile No",width=20,pady=10)
lblMobileNo.grid(row=5,column=0)
txtMobileNo = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=MobileNo)
txtMobileNo.grid(row=5,column=1)

lblResidence = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Residence",width=20,pady=10)
lblResidence.grid(row=5,column=2)
txtResidence = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=Residence)
txtResidence.grid(row=5,column=3)

lblCaste = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Caste",width=20,pady=10)
lblCaste.grid(row=6,column=0)
cboCaste = ttk.Combobox(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=Caste)
cboCaste.grid(row=6,column=1)
cboCaste['values']=('','General/OC','BC','SC','ST')
cboCaste.current(0)

lblSubCaste = Label(AddNewFrame1,font=('arial',12,'bold'),text = "SubCaste",width=20,pady=10)
lblSubCaste.grid(row=6,column=2)
txtSubCaste = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=SubCaste)
txtSubCaste.grid(row=6,column=3)

lblRationCardNo = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Ration Card No",width=20,pady=10)
lblRationCardNo.grid(row=7,column=0)
txtRationCardNo = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=RationCardNo)
txtRationCardNo.grid(row=7,column=1)

lblBankAccountNo = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Bank Account No",width=20,pady=10)
lblBankAccountNo.grid(row=7,column=2)
txtBankAccountNo = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=BankAccountNo)
txtBankAccountNo.grid(row=7,column=3)

lblIfscCode = Label(AddNewFrame1,font=('arial',12,'bold'),text = "IFSC Code",width=20,pady=10)
lblIfscCode.grid(row=8,column=0)
txtIfscCode = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=IfscCode)
txtIfscCode.grid(row=8,column=1)

lblBankName = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Bank Name",width=20,pady=10)
lblBankName.grid(row=8,column=2)
txtBankName = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=BankName)
txtBankName.grid(row=8,column=3)

lblClass = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Class",width=20,pady=10)
lblClass.grid(row=9,column=0)
txtClass = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left',textvariable=Class)
txtClass.grid(row=9,column=1)

lblAdmissionNo = Label(AddNewFrame1,font=('arial',12,'bold'),text = "Admission No",width=20,pady=10)
lblAdmissionNo.grid(row=9,column=2)
txtAdmissionNo = Entry(AddNewFrame1,font=('arial',12,'bold'),width=35,justify='left')
txtAdmissionNo.grid(row=9,column=3)

btnAddData = Button(AddNewFrame2,font=('arial',12,'bold'),text="Add Data",padx=5,pady=1,command=AddData,
                    width=10).grid(row=0,column=0)
btnResetData = Button(AddNewFrame2,font=('arial',12,'bold'),text="Reset Data",padx=5,pady=1,command=Reset,
                    width=10).grid(row=0,column=1)

#====================================================================================================================

lblAdmissionNo = Label(SearchFrame1,font=('arial',12,'bold'),text = "Admission No",width=20,pady=10)
lblAdmissionNo.grid(row=0,column=0)
txtAdmissionNo = Entry(SearchFrame1,font=('arial',12,'bold'),width=36,justify='left',textvariable=AdmissionNo)
txtAdmissionNo.grid(row=0,column=1)

lblStudentName = Label(SearchFrame1,font=('arial',12,'bold'),text = "Student Name",width=20,pady=10)
lblStudentName.grid(row=0,column=2)
txtStudentName = Entry(SearchFrame1,font=('arial',12,'bold'),width=36,justify='left',textvariable=StudentName)
txtStudentName.grid(row=0,column=3)

lblChildId = Label(SearchFrame2,font=('arial',12,'bold'),text = "Child Id",width=20,pady=10)
lblChildId.grid(row=1,column=0)
txtChildId = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=ChildId)
txtChildId.grid(row=1,column=1)

lblGender = Label(SearchFrame2,font=('arial',12,'bold'),text = "Gender",width=20,pady=10)
lblGender.grid(row=1,column=2)
cboGender = ttk.Combobox(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=Gender)
cboGender.grid(row=1,column=3)
cboGender['values'] = ('','Girl','Boy')
cboGender.current(0)

lblDateOfBirth = Label(SearchFrame2,font=('arial',12,'bold'),text = "Date Of Birth(dd/mm/yy)",width=20,pady=10)
lblDateOfBirth.grid(row=2,column=0)
txtDateOfBirth = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=DateOfBirth)
txtDateOfBirth.grid(row=2,column=1)

lblDateOfJoining = Label(SearchFrame2,font=('arial',12,'bold'),text = "Date Of Joining(dd/mm/yy)",width=20,pady=10)
lblDateOfJoining.grid(row=2,column=2)
txtDateOfJoining = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=DateOfJoining)
txtDateOfJoining.grid(row=2,column=3)

lblAadharNo = Label(SearchFrame2,font=('arial',12,'bold'),text = "Aadhar Card No",width=20,pady=10)
lblAadharNo.grid(row=3,column=0)
txtAadharNo = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=AadharNo)
txtAadharNo.grid(row=3,column=1)

lblFatherName = Label(SearchFrame2,font=('arial',12,'bold'),text = "Father Name",width=20,pady=10)
lblFatherName.grid(row=3,column=2)
txtFatherName = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=FatherName)
txtFatherName.grid(row=3,column=3)

lblMotherName = Label(SearchFrame2,font=('arial',12,'bold'),text = "Mother Name",width=20,pady=10)
lblMotherName.grid(row=4,column=0)
txtMotherName = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=MotherName)
txtMotherName.grid(row=4,column=1)

lblGuardianName = Label(SearchFrame2,font=('arial',12,'bold'),text = "Guardian Name",width=20,pady=10)
lblGuardianName.grid(row=4,column=2)
txtGuardianName = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=GuardianName)
txtGuardianName.grid(row=4,column=3)

lblMobileNo = Label(SearchFrame2,font=('arial',12,'bold'),text = "Mobile No",width=20,pady=10)
lblMobileNo.grid(row=5,column=0)
txtMobileNo = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=MobileNo)
txtMobileNo.grid(row=5,column=1)

lblResidence = Label(SearchFrame2,font=('arial',12,'bold'),text = "Residence",width=20,pady=10)
lblResidence.grid(row=5,column=2)
txtResidence = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=Residence)
txtResidence.grid(row=5,column=3)

lblCaste = Label(SearchFrame2,font=('arial',12,'bold'),text = "Caste",width=20,pady=10)
lblCaste.grid(row=6,column=0)
cboCaste = ttk.Combobox(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=Caste)
cboCaste.grid(row=6,column=1)
cboCaste['values']=('','General/OC','BC','SC','ST')
cboCaste.current(0)

lblSubCaste = Label(SearchFrame2,font=('arial',12,'bold'),text = "SubCaste",width=20,pady=10)
lblSubCaste.grid(row=6,column=2)
txtSubCaste = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=SubCaste)
txtSubCaste.grid(row=6,column=3)

lblRationCardNo = Label(SearchFrame2,font=('arial',12,'bold'),text = "Ration Card No",width=20,pady=10)
lblRationCardNo.grid(row=7,column=0)
txtRationCardNo = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=RationCardNo)
txtRationCardNo.grid(row=7,column=1)

lblBankAccountNo = Label(SearchFrame2,font=('arial',12,'bold'),text = "Bank Account No",width=20,pady=10)
lblBankAccountNo.grid(row=7,column=2)
txtBankAccountNo = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=BankAccountNo)
txtBankAccountNo.grid(row=7,column=3)

lblIfscCode = Label(SearchFrame2,font=('arial',12,'bold'),text = "IFSC Code",width=20,pady=10)
lblIfscCode.grid(row=8,column=0)
txtIfscCode = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=IfscCode)
txtIfscCode.grid(row=8,column=1)

lblBankName = Label(SearchFrame2,font=('arial',12,'bold'),text = "Bank Name",width=20,pady=10)
lblBankName.grid(row=8,column=2)
txtBankName = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=BankName)
txtBankName.grid(row=8,column=3)

lblClass = Label(SearchFrame2,font=('arial',12,'bold'),text = "Class",width=20,pady=10)
lblClass.grid(row=9,column=0)
txtClass = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left',textvariable=Class)
txtClass.grid(row=9,column=1)

lblAdmissionNo = Label(SearchFrame2,font=('arial',12,'bold'),text = "Admission No",width=20,pady=10)
lblAdmissionNo.grid(row=9,column=2)
txtAdmissionNo = Entry(SearchFrame2,font=('arial',12,'bold'),width=35,justify='left')
txtAdmissionNo.grid(row=9,column=3)

btnSearchData = Button(SearchFrame3,font=('arial',12,'bold'),text="Search Data",padx=5,pady=1,command = SearchData,
                    width=10).grid(row=0,column=0)
btnDeleteData = Button(SearchFrame3,font=('arial',12,'bold'),text="Delete Data",padx=5,pady=1,command = DeleteData,
                    width=10).grid(row=0,column=1)
btnUpdateData = Button(SearchFrame3,font=('arial',12,'bold'),text="Update Data",padx=5,pady=1,command = UpdateData,
                    width=10).grid(row=0,column=2)
btnResetData = Button(SearchFrame3,font=('arial',12,'bold'),text="Reset Data",padx=5,pady=1,command = Reset,
                    width=10).grid(row=0,column=3)

#====================================================================================================================

win.mainloop()
